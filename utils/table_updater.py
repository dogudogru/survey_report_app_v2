import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from config.constants import PARTY_MAPPING
from datetime import datetime
import calendar
import os

class BaseTableUpdater:
    def __init__(self, template_path: str, output_path: str):
        print(f"Initializing TableUpdater with template: {template_path}, output: {output_path}")
        self.template_path = template_path
        self.output_path = output_path
        self.workbook = None
        
        # Turkish month abbreviations
        self.tr_months = {
            1: 'Oca', 2: 'Şub', 3: 'Mar', 4: 'Nis', 5: 'May', 6: 'Haz',
            7: 'Tem', 8: 'Ağu', 9: 'Eyl', 10: 'Eki', 11: 'Kas', 12: 'Ara'
        }
        
        # Common mappings
        self.party_mapping_2023 = {
            'Adalet ve Kalkınma Partisi (AK Parti/AKP)': 'AK Parti',
            'Cumhuriyet Halk Partisi (CHP)': 'CHP',
            'Yeşil Sol Parti (YSP) / Halkların Demokratik Partisi (HDP) / DEM Parti': 'Yeşil Sol Parti',
            'İYİ Parti': 'İYİ Parti',
            'Milliyetçi Hareket Partisi (MHP)': 'MHP'
        }
        
        self.econ_current_mapping = {
            'Çok kötü': 'Çok kötü',
            'Kötü': 'Kötü',
            'Ne iyi ne kötü': 'Ne iyi ne kötü',
            'İyi': 'İyi',
            'Çok iyi': 'Çok iyi'
        }
        
        self.econ_future_mapping = {
            'Çok daha kötü': 'Çok daha kötü',
            'Daha kötü': 'Daha kötü',
            'Değişmez': 'Değişmez',
            'Daha iyi': 'Daha iyi',
            'Çok daha iyi': 'Çok daha iyi'
        }

        # English translation mappings
        self.en_months = {
            1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
            7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
        }

        self.en_party_mapping = {
            'AK Parti': 'AK Party',
            'CHP': 'CHP',
            'DEM Parti': 'HDP/Green Left',
            'Yeşil Sol Parti': 'HDP/Green Left',
            'DEM Parti/YSP': 'HDP/Green Left',
            'DEM Parti / YSP': 'HDP/Green Left',
            'İYİ Parti': 'İYİ Party',
            'MHP': 'MHP',
            'Kararsız': 'Undecided',
            'Kararsızım': 'Undecided',
            'Oy Kullanmam': '''Won't Vote''',
            'Oy kullanmayacağım': '''Won't Vote''',
            'Diğer': 'Other',
            'Zafer Partisi': 'Victory Party',
            'Yeniden Refah': 'New Welfare Party',
            'Anahtar Parti': 'Key Party',
            'Toplam': 'Total',
            '2023 Milletvekili Seçimleri Oy Tercihi': '2023 Parliamentary Election Vote Choice'
        }

        self.en_econ_current_mapping = {
            'Çok kötü': 'Very Bad',
            'Kötü': 'Bad',
            'Ne iyi ne kötü': 'Neither Good nor Bad',
            'İyi': 'Good',
            'Çok iyi': 'Very Good'
        }

        self.en_econ_future_mapping = {
            'Çok daha kötü': 'Much Worse',
            'Daha kötü': 'Worse',
            'Değişmez': 'No Change',
            'Daha iyi': 'Better',
            'Çok daha iyi': 'Much Better'
        }

        self.en_education_mapping = {
            'İlköğretim ve altı': 'Primary School or Less',
            'Lise': 'High School',
            'Yüksekokul ve üzeri': 'University or Higher'
        }

        self.en_job_mapping = {
            'Emekli, çalışmıyor': 'Retired, Not Working',
            'Çalışmıyor,  iş aramıyor': 'Unemployed, Not Looking',
            'İşsiz': 'Unemployed, Looking',
            'Kendi hesabına çalışan veya işveren': 'Self-Employed or Employer',
            'Maaşlı devlet çalışanı': 'Public Sector Employee',
            'Öğrenci': 'Student',
            'Ücretli özel sektör çalışanı': 'Private Sector Employee',
            'Günlük/yevmiyeli çalışan': 'Daily Wage Worker',
            'Ücretli ya da yevmiyeli çalışan': 'Daily Wage Worker'
        }

        self.en_subsistence_mapping = {
            'Geçtiğimiz ay gelirim giderlerimi karşılamadı.': 'My income did not meet my expenses.',
            'Geçtiğimiz ay gelirim giderlerimi ucu ucuna karşıladı.': 'My income barely met my expenses.',
            'Geçtiğimiz ay gelirim giderlerimin üzerinde oldu.': 'My income exceeded my expenses.',
            'Geçtiğimiz ay gelirim giderlerimi fazlasıyla karşıladı.': 'My income was well above my expenses.',
            'Kadın': 'Woman',
            'Erkek': 'Man'
        }

        self.en_months_mapping = {
            'Oca.': 'Jan.',
            'Şub.': 'Feb.',
            'Mar.': 'Mar.',
            'Nis.': 'Apr.',
            'May.': 'May.',
            'Haz.': 'Jun.',
            'Tem.': 'Jul.',
            'Ağu.': 'Aug.',
            'Eyl.': 'Sep.',
            'Eki.': 'Oct.',
            'Kas.': 'Nov.',
            'Ara.': 'Dec.'
        }
    
    def _load_workbook(self):
        """Load the workbook if not already loaded"""
        if self.workbook is None:
            try:
                print(f"\nLoading workbook from: {self.template_path}")
                if not os.path.exists(self.template_path):
                    raise Exception(f"Template file does not exist: {self.template_path}")
                
                print(f"File exists, size: {os.path.getsize(self.template_path)} bytes")
                
                # Try to read the first few bytes to check if file is accessible
                with open(self.template_path, 'rb') as f:
                    first_bytes = f.read(10)
                    print(f"First few bytes: {first_bytes}")
                
                print("Loading workbook with openpyxl...")
                self.workbook = load_workbook(self.template_path)
                
                if self.workbook is None:
                    raise Exception("load_workbook returned None")
                    
                print(f"Successfully loaded workbook. Type: {type(self.workbook)}")
                print(f"Available sheets: {self.workbook.sheetnames}")
                print("Sheet details:")
                for sheet in self.workbook.sheetnames:
                    print(f"  - {sheet}")
            except Exception as e:
                print(f"Error loading workbook: {str(e)}")
                print(f"Error type: {type(e)}")
                import traceback
                print(f"Traceback: {traceback.format_exc()}")
                raise Exception(f"Error loading workbook: {str(e)}")
        else:
            print("Workbook already loaded")
    
    def _save_workbook(self):
        """Save and close the workbook"""
        if self.workbook is not None:
            try:
                print(f"Saving workbook to: {self.output_path}")
                self.workbook.save(self.output_path)
                print(f"Successfully saved workbook, size: {os.path.getsize(self.output_path)} bytes")
                self.workbook = None
            except Exception as e:
                print(f"Error saving workbook: {str(e)}")
                raise Exception(f"Error saving workbook: {str(e)}")
    
    def _get_worksheet(self, sheet_name: str):
        """Safely get a worksheet by name"""
        print(f"\nAttempting to get worksheet: {sheet_name}")
        
        if self.workbook is None:
            print("Error: Workbook is None. Checking if _load_workbook was called...")
            self._load_workbook()
            if self.workbook is None:
                raise Exception("Workbook is still None after loading")
        
        print(f"Workbook loaded. Type: {type(self.workbook)}")
        print(f"Available sheets: {self.workbook.sheetnames}")
        
        # Check if sheet name exists (case-sensitive)
        if sheet_name not in self.workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found.")
            print("Available sheets (with string representation):")
            for sheet in self.workbook.sheetnames:
                print(f"  - '{sheet}' (type: {type(sheet)})")
            raise Exception(f"Sheet '{sheet_name}' not found. Available sheets: {self.workbook.sheetnames}")
        
        print(f"Found sheet: {sheet_name}")
        worksheet = self.workbook[sheet_name]
        print(f"Successfully retrieved worksheet: {worksheet}")
        return worksheet
    
    def _find_column(self, df: pd.DataFrame, search_text: str) -> str:
        """Find the exact column name that contains the given text"""
        matching_cols = [col for col in df.columns if search_text in col]
        if matching_cols:
            return matching_cols[0]
        raise ValueError(f"Could not find column containing: {search_text}")
    
    def _create_pivot_table(self, df: pd.DataFrame, values: str, index: str, columns: str, 
                          aggfunc: str = 'sum', calc_method: str = 'percent_of_column') -> pd.DataFrame:
        """Create a pivot table with specified calculation method"""
        try:
            pivot = pd.pivot_table(
                df,
                values=values,
                index=index,
                columns=columns,
                aggfunc=aggfunc
            )
            
            if calc_method == 'percent_of_column':
                pivot_pct = pivot.div(pivot.sum(axis=0), axis=1) * 100
            elif calc_method == 'percent_of_row':
                pivot_pct = pivot.div(pivot.sum(axis=1), axis=0) * 100
            else:
                raise ValueError(f"Unknown calculation method: {calc_method}")
                
            return pivot_pct
        except Exception as e:
            raise Exception(f"Error creating pivot table: {str(e)}")
    
    def _apply_conditional_formatting(self, worksheet, cell_range: str, color_scale: str = 'white_to_plum'):
        """Apply conditional formatting to specified range"""
        if color_scale == 'white_to_plum':
            end_color = 'DDA0DD'  # Plum
        elif color_scale == 'white_to_indigo':
            end_color = '4B0082'  # Indigo
        else:
            raise ValueError(f"Unknown color scale: {color_scale}")
            
        color_scale_rule = ColorScaleRule(
            start_type='min',
            start_color='FFFFFF',  # White
            end_type='max',
            end_color=end_color
        )
        
        worksheet.conditional_formatting.add(cell_range, color_scale_rule)
    
    def _get_current_month_str(self) -> str:
        """Get current month in specified language format (e.g., 'Oca.24' or 'Jan.24')"""
        current_date = datetime.now()
        if self.language == 'tr':
            return f"{self.tr_months[current_date.month]}.{str(current_date.year)[2:]}"
        else:
            return f"{self.en_months[current_date.month]}.{str(current_date.year)[2:]}"
    
    def _round_values(self, value: float) -> int:
        """Round values to whole numbers between 0 and 100"""
        if pd.isna(value) or not np.isfinite(value):
            return 0
        return min(max(round(value), 0), 100)
    
    def _update_cell_value(self, worksheet, cell: str, value: float):
        """Update cell with rounded value"""
        worksheet[cell] = self._round_values(value)
    
    def _shift_historical_data(self, ws, start_row: int, end_row: int):
        """Shift historical data left by one column (G to K) with proper month translation"""
        for col_idx in range(ord('G'), ord('L')):
            col = chr(col_idx)
            next_col = chr(col_idx + 1)
            
            # Copy and translate header if needed
            header_value = ws[f'{next_col}1'].value
            if self.language == 'en' and header_value:
                # Split the header value into month and year (e.g., "Nis.23" -> ["Nis", "23"])
                try:
                    month_part = header_value.split('.')[0] + '.'  # Add the dot back
                    year_part = header_value.split('.')[1]
                    
                    # If the month part exists in our mapping, translate it and keep the year
                    if month_part in self.en_months_mapping:
                        ws[f'{col}1'] = f"{self.en_months_mapping[month_part]}{year_part}"
                    else:
                        ws[f'{col}1'] = header_value
                except:
                    # If splitting fails, keep original value
                    ws[f'{col}1'] = header_value
            else:
                ws[f'{col}1'] = header_value
            
            # Copy values
            for row in range(start_row, end_row + 1):
                ws[f'{col}{row}'] = ws[f'{next_col}{row}'].value

class TableUpdater(BaseTableUpdater):
    def __init__(self, template_path: str, output_path: str, language: str = 'tr'):
        super().__init__(template_path, output_path)
        self.language = language
    
    def _get_current_month_str(self) -> str:
        """Get current month in specified language format (e.g., 'Oca.24' or 'Jan.24')"""
        current_date = datetime.now()
        if self.language == 'tr':
            return f"{self.tr_months[current_date.month]}.{str(current_date.year)[2:]}"
        else:
            return f"{self.en_months[current_date.month]}.{str(current_date.year)[2:]}"
    
    def _translate_worksheet_text(self, ws):
        """Translate worksheet text from Turkish to English"""
        if self.language == 'tr':
            return  # No translation needed for Turkish
            
        # Get all cells in the worksheet
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    value = str(cell.value)
                    # Try different mappings
                    if value in self.en_party_mapping:
                        cell.value = self.en_party_mapping[value]
                    elif value in self.en_econ_current_mapping:
                        cell.value = self.en_econ_current_mapping[value]
                    elif value in self.en_econ_future_mapping:
                        cell.value = self.en_econ_future_mapping[value]
                    elif value in self.en_education_mapping:
                        cell.value = self.en_education_mapping[value]
                    elif value in self.en_job_mapping:
                        cell.value = self.en_job_mapping[value]
                    elif value in self.en_subsistence_mapping:
                        cell.value = self.en_subsistence_mapping[value]
                    elif value in self.en_months_mapping:
                        cell.value = self.en_months_mapping[value]
    

    def update_all_tables(self, survey_data: pd.DataFrame, historical_data: pd.DataFrame = None):
        """Update all tables in the workbook"""
        try:
            self._load_workbook()
            
            # Update each table
            print("Updating 2023 party table...")
            self.update_2023_party_table(survey_data)
            
            print("Updating economic current party table...")
            self.update_econ_current_party_table(survey_data)
            
            print("Updating economic current age table...")
            self.update_econ_current_age_table(survey_data)
            
            print("Updating economic current education table...")
            self.update_econ_current_education_table(survey_data)
            
            print("Updating economic current jobs table...")
            self.update_econ_current_jobs_table(survey_data)
            
            print("Updating economic future party table...")
            self.update_econ_future_party_table(survey_data)
            
            print("Updating economic future age table...")
            self.update_econ_future_age_table(survey_data)
            
            print("Updating economic future jobs table...")
            self.update_econ_future_jobs_table(survey_data)
            
            print("Updating economic current vs future table...")
            self.update_econ_current_vs_future_table(survey_data)
            
            print("Updating subsistence demographics table...")
            self.update_subsistence_demographics_table(survey_data)
            
            print("Updating subsistence party education table...")
            self.update_subsistence_party_education_table(survey_data)
            
            print("Updating subsistence jobs table...")
            self.update_subsistence_jobs_table(survey_data)
            
            # Translate worksheet text if English
            if self.language == 'en':
                print("Translating worksheets to English...")
                for sheet_name in self.workbook.sheetnames:
                    self._translate_worksheet_text(self.workbook[sheet_name])
            
            # Save the workbook
            self._save_workbook()
        except Exception as e:
            raise Exception(f"Error updating tables: {str(e)}")
    
    def update_2023_party_table(self, survey_data: pd.DataFrame):
        """Update the 2023 party transition table (27_party_2023)"""
        try:
            # Map current and 2023 party names
            survey_data['mapped_party'] = survey_data['parti'].map(PARTY_MAPPING)
            survey_data['mapped_2023_party'] = survey_data['2023 Genel Seçimlerinde hangi partiye oy verdiniz?'].map(self.party_mapping_2023)
            
            # Define valid parties
            valid_parties = ['AK Parti', 'CHP', 'İYİ Parti', 'DEM Parti', 'MHP', 
                            'Yeniden Refah Partisi', 'Zafer Partisi', 
                            'Anahtar Parti', 'Oy kullanmayacağım', 'Kararsızım']
            
            # Map non-valid parties to 'Diğer'
            survey_data['mapped_party'] = survey_data['mapped_party'].apply(
                lambda x: x if x in valid_parties else 'Diğer'
            )
            
            # Filter for relevant 2023 parties
            relevant_parties = ['AK Parti', 'CHP', 'MHP', 'İYİ Parti', 'Yeşil Sol Parti']
            filtered_data = survey_data[survey_data['mapped_2023_party'].isin(relevant_parties)]
            
            # Create pivot table
            pivot_pct = self._create_pivot_table(
                filtered_data,
                values='duzeltilmis_agirlik',
                index='mapped_party',
                columns='mapped_2023_party',
                calc_method='percent_of_column'
            )
            
            # Get worksheet
            ws = self._get_worksheet('27_party_2023')
            print("Get 27_party_2023")
            
            # Update values in B3:F14
            row_mapping = {
                'AK Parti': 3,
                'CHP': 4,
                'MHP': 5,
                'İYİ Parti': 6,
                'DEM Parti': 7,
                'Yeniden Refah Partisi': 8,
                'Zafer Partisi': 9,
                'Anahtar Parti': 10,
                'Diğer': 11,
                'Oy kullanmayacağım': 12,
                'Kararsızım': 13,
                'Toplam': 14
            }
            
            col_mapping = {
                'AK Parti': 'B',
                'CHP': 'C',
                'MHP': 'D',
                'İYİ Parti': 'E',
                'Yeşil Sol Parti': 'F'
            }
            
            # Update each cell with the new percentage
            for current_party, row_num in row_mapping.items():
                for party_2023, col_letter in col_mapping.items():
                    cell_coord = f'{col_letter}{row_num}'
                    if current_party == 'Toplam':
                        self._update_cell_value(ws, cell_coord, 100)
                    else:
                        try:
                            value = pivot_pct.loc[current_party, party_2023]
                            self._update_cell_value(ws, cell_coord, value)
                        except KeyError:
                            self._update_cell_value(ws, cell_coord, 0)
            
            # Apply conditional formatting
            #self._apply_conditional_formatting(ws, 'B3:F13', 'white_to_plum')
        except Exception as e:
            raise Exception(f"Error updating 2023 party table: {str(e)}")
    
    def update_econ_current_party_table(self, survey_data: pd.DataFrame):
        """Update the economic situation by party table (34_econ_current_party)"""
        try:
            # Find the economy question column
            econ_col = self._find_column(survey_data, "Bugün itibari ile ekonominin nasıl olduğunu düşünüyorsunuz")
            
            # Map party names and responses
            survey_data['mapped_party'] = survey_data['2023 Genel Seçimlerinde hangi partiye oy verdiniz?'].map(self.party_mapping_2023)
            survey_data['econ_response'] = survey_data[econ_col].map(self.econ_current_mapping)
            
            # Create pivot table
            pivot_pct = self._create_pivot_table(
                survey_data,
                values='duzeltilmis_agirlik',
                index='mapped_party',
                columns='econ_response',
                calc_method='percent_of_row'
            )
            
            # Get worksheet
            ws = self._get_worksheet('34_econ_current_party')
            print("Get 34_econ_current_party")

            # Update current month values (B to F columns)
            row_mapping = {
                'AK Parti': 2,
                'CHP': 3,
                'Yeşil Sol Parti': 4,
                'İYİ Parti': 5,
                'MHP': 6
            }
            
            col_mapping = {
                'Çok kötü': 'B',
                'Kötü': 'C',
                'Ne iyi ne kötü': 'D',
                'İyi': 'E',
                'Çok iyi': 'F'
            }
            
            # Update each cell with the new percentage
            for party, row_num in row_mapping.items():
                for response, col_letter in col_mapping.items():
                    cell_coord = f'{col_letter}{row_num}'
                    try:
                        value = pivot_pct.loc[party, response]
                        self._update_cell_value(ws, cell_coord, value)
                    except KeyError:
                        self._update_cell_value(ws, cell_coord, 0)
            
            # Calculate negative percentages for current month
            negative_pcts = {}
            for party in row_mapping.keys():
                try:
                    negative_pct = pivot_pct.loc[party, 'Çok kötü'] + pivot_pct.loc[party, 'Kötü']
                    negative_pcts[party] = self._round_values(negative_pct)
                except KeyError:
                    negative_pcts[party] = 0
            
            # Shift historical data left by one column (G to K)
            self._shift_historical_data(ws, 2, 6)
            
            # Update the last column (L) with current month's data
            current_month = self._get_current_month_str()
            ws['L1'] = current_month
            
            for party, row_num in row_mapping.items():
                self._update_cell_value(ws, f'L{row_num}', negative_pcts[party])
            
            # Apply conditional formatting
            #self._apply_conditional_formatting(ws, 'B2:F6', 'white_to_plum')  # Current month data
            #self._apply_conditional_formatting(ws, 'G2:L6', 'white_to_indigo')  # Historical data
        except Exception as e:
            raise Exception(f"Error updating economic current party table: {str(e)}")
    
    def update_econ_current_age_table(self, survey_data: pd.DataFrame):
        """Update the economic situation by age table (36_econ_current_age)"""
        try:
            # Find the economy question column
            econ_col = self._find_column(survey_data, "Bugün itibari ile ekonominin nasıl olduğunu düşünüyorsunuz")
            
            # Map responses
            survey_data['econ_response'] = survey_data[econ_col].map(self.econ_current_mapping)
            survey_data['Yaş grubu'] = survey_data['Yaş grubu'].replace('65 ve üstü', '65+')
            
            # Create pivot table
            pivot_pct = self._create_pivot_table(
                survey_data,
                values='duzeltilmis_agirlik',
                index='Yaş grubu',
                columns='econ_response',
                calc_method='percent_of_row'
            )
            
            # Get worksheet
            ws = self._get_worksheet('36_econ_current_age')
            print("Get 36_econ_current_age")
            
            # Update current month values (B to F columns)
            row_mapping = {
                '18-24': 2,
                '25-34': 3,
                '35-44': 4,
                '45-54': 5,
                '55-64': 6,
                '65+': 7
            }
            
            col_mapping = {
                'Çok kötü': 'B',
                'Kötü': 'C',
                'Ne iyi ne kötü': 'D',
                'İyi': 'E',
                'Çok iyi': 'F'
            }
            
            # Update each cell with the new percentage
            for age_group, row_num in row_mapping.items():
                for response, col_letter in col_mapping.items():
                    cell_coord = f'{col_letter}{row_num}'
                    try:
                        value = pivot_pct.loc[age_group, response]
                        self._update_cell_value(ws, cell_coord, value)
                    except KeyError:
                        self._update_cell_value(ws, cell_coord, 0)
            
            # Calculate negative percentages for current month
            negative_pcts = {}
            for age_group in row_mapping.keys():
                try:
                    negative_pct = pivot_pct.loc[age_group, 'Çok kötü'] + pivot_pct.loc[age_group, 'Kötü']
                    negative_pcts[age_group] = self._round_values(negative_pct)
                except KeyError:
                    negative_pcts[age_group] = 0
            
            # Shift historical data left by one column (G to K)
            self._shift_historical_data(ws, 2, 7)
            
            # Update the last column (L) with current month's data
            current_month = self._get_current_month_str()
            ws['L1'] = current_month
            
            for age_group, row_num in row_mapping.items():
                self._update_cell_value(ws, f'L{row_num}', negative_pcts[age_group])
            
            # Apply conditional formatting
            #self._apply_conditional_formatting(ws, 'B2:F7', 'white_to_plum')  # Current month data
            #self._apply_conditional_formatting(ws, 'G2:L7', 'white_to_indigo')  # Historical data
        except Exception as e:
            raise Exception(f"Error updating economic current age table: {str(e)}")
    
    def update_econ_current_education_table(self, survey_data: pd.DataFrame):
        """Update the economic situation by education table (38_econ_current_education)"""
        try:
            # Find the economy question column
            econ_col = self._find_column(survey_data, "Bugün itibari ile ekonominin nasıl olduğunu düşünüyorsunuz")
            
            # Map responses
            survey_data['econ_response'] = survey_data[econ_col].map(self.econ_current_mapping)
            
            # Create pivot table
            pivot_pct = self._create_pivot_table(
                survey_data,
                values='duzeltilmis_agirlik',
                index='education',
                columns='econ_response',
                calc_method='percent_of_row'
            )
            
            # Get worksheet
            ws = self._get_worksheet('38_econ_current_education')
            print("Get 38_econ_current_education")
            
            # Update current month values (B to F columns)
            row_mapping = {
                'İlköğretim ve altı': 2,
                'Lise': 3,
                'Yüksekokul ve üzeri': 4
            }
            
            col_mapping = {
                'Çok kötü': 'B',
                'Kötü': 'C',
                'Ne iyi ne kötü': 'D',
                'İyi': 'E',
                'Çok iyi': 'F'
            }
            
            # Update each cell with the new percentage
            for edu_level, row_num in row_mapping.items():
                for response, col_letter in col_mapping.items():
                    cell_coord = f'{col_letter}{row_num}'
                    try:
                        value = pivot_pct.loc[edu_level, response]
                        self._update_cell_value(ws, cell_coord, value)
                    except KeyError:
                        self._update_cell_value(ws, cell_coord, 0)
            
            # Calculate negative percentages for current month
            negative_pcts = {}
            for edu_level in row_mapping.keys():
                try:
                    negative_pct = pivot_pct.loc[edu_level, 'Çok kötü'] + pivot_pct.loc[edu_level, 'Kötü']
                    negative_pcts[edu_level] = self._round_values(negative_pct)
                except KeyError:
                    negative_pcts[edu_level] = 0
            
            # Shift historical data left by one column (G to K)
            self._shift_historical_data(ws, 2, 4)
            
            # Update the last column (L) with current month's data
            current_month = self._get_current_month_str()
            ws['L1'] = current_month
            
            for edu_level, row_num in row_mapping.items():
                self._update_cell_value(ws, f'L{row_num}', negative_pcts[edu_level])
            
            # Apply conditional formatting
            #self._apply_conditional_formatting(ws, 'B2:F4', 'white_to_plum')  # Current month data
            #self._apply_conditional_formatting(ws, 'G2:L4', 'white_to_indigo')  # Historical data
        except Exception as e:
            raise Exception(f"Error updating economic current education table: {str(e)}")
    
    def update_econ_future_party_table(self, survey_data: pd.DataFrame):
        """Update the future economic situation by party table (42_econ_future_party)"""
        try:
            # Find the economy question column
            econ_col = self._find_column(survey_data, "Önümüzdeki bir yıl içerisinde ekonominin nasıl olacağını düşünüyorsunuz")
            print("found econ_col")
            
            # Map party names and responses
            survey_data['mapped_party'] = survey_data['2023 Genel Seçimlerinde hangi partiye oy verdiniz?'].map(self.party_mapping_2023)
            survey_data['econ_response'] = survey_data[econ_col].map(self.econ_future_mapping)
            
            # Create pivot table
            pivot_pct = self._create_pivot_table(
                survey_data,
                values='duzeltilmis_agirlik',
                index='mapped_party',
                columns='econ_response',
                calc_method='percent_of_row'
            )
            
            # Get worksheet
            ws = self._get_worksheet('42_econ_future_party')
            print("Get 42_econ_future_party")
            print(pivot_pct)
            
            # Update current month values (B to F columns)
            row_mapping = {
                'AK Parti': 2,
                'CHP': 3,
                'Yeşil Sol Parti': 4,
                'İYİ Parti': 5,
                'MHP': 6
            }
            
            col_mapping = {
                'Çok daha kötü': 'B',
                'Daha kötü': 'C',
                'Değişmez': 'D',
                'Daha iyi': 'E',
                'Çok daha iyi': 'F'
            }
            
            # Update each cell with the new percentage
            for party, row_num in row_mapping.items():
                for response, col_letter in col_mapping.items():
                    cell_coord = f'{col_letter}{row_num}'
                    try:
                        value = pivot_pct.loc[party, response]
                        self._update_cell_value(ws, cell_coord, value)
                    except KeyError:
                        self._update_cell_value(ws, cell_coord, 0)
            
            # Calculate negative percentages for current month
            negative_pcts = {}
            for party in row_mapping.keys():
                try:
                    negative_pct = pivot_pct.loc[party, 'Çok daha kötü'] + pivot_pct.loc[party, 'Daha kötü']
                    negative_pcts[party] = self._round_values(negative_pct)
                except KeyError:
                    negative_pcts[party] = 0
            
            # Shift historical data left by one column (G to K)
            self._shift_historical_data(ws, 2, 6)
            
            # Update the last column (L) with current month's data
            current_month = self._get_current_month_str()
            ws['L1'] = current_month
            
            for party, row_num in row_mapping.items():
                self._update_cell_value(ws, f'L{row_num}', negative_pcts[party])
            
            # Apply conditional formatting
            #self._apply_conditional_formatting(ws, 'B2:F6', 'white_to_plum')  # Current month data
            #self._apply_conditional_formatting(ws, 'G2:L6', 'white_to_indigo')  # Historical data
        except Exception as e:
            raise Exception(f"Error updating economic future party table: {str(e)}")
    
    def update_econ_future_age_table(self, survey_data: pd.DataFrame):
        """Update the future economic situation by age table (44_econ_future_age)"""
        try:
            # Find the economy question column
            econ_col = self._find_column(survey_data, "Önümüzdeki bir yıl içerisinde ekonominin nasıl olacağını düşünüyorsunuz")
            
            # Map responses
            survey_data['econ_response'] = survey_data[econ_col].map(self.econ_future_mapping)
            survey_data['Yaş grubu'] = survey_data['Yaş grubu'].replace('65 ve üstü', '65+')
            
            # Create pivot table
            pivot_pct = self._create_pivot_table(
                survey_data,
                values='duzeltilmis_agirlik',
                index='Yaş grubu',
                columns='econ_response',
                calc_method='percent_of_row'
            )
            
            # Get worksheet
            ws = self._get_worksheet('44_econ_future_age')
            
            # Update current month values (B to F columns)
            row_mapping = {
                '18-24': 2,
                '25-34': 3,
                '35-44': 4,
                '45-54': 5,
                '55-64': 6,
                '65+': 7
            }
            
            col_mapping = {
                'Çok daha kötü': 'B',
                'Daha kötü': 'C',
                'Değişmez': 'D',
                'Daha iyi': 'E',
                'Çok daha iyi': 'F'
            }
            
            # Update each cell with the new percentage
            for age_group, row_num in row_mapping.items():
                for response, col_letter in col_mapping.items():
                    cell_coord = f'{col_letter}{row_num}'
                    try:
                        value = pivot_pct.loc[age_group, response]
                        self._update_cell_value(ws, cell_coord, value)
                    except KeyError:
                        self._update_cell_value(ws, cell_coord, 0)
            
            # Calculate negative percentages for current month
            negative_pcts = {}
            for age_group in row_mapping.keys():
                try:
                    negative_pct = pivot_pct.loc[age_group, 'Çok daha kötü'] + pivot_pct.loc[age_group, 'Daha kötü']
                    negative_pcts[age_group] = self._round_values(negative_pct)
                except KeyError:
                    negative_pcts[age_group] = 0
            
            # Shift historical data left by one column (G to K)
            self._shift_historical_data(ws, 2, 7)
            
            # Update the last column (L) with current month's data
            current_month = self._get_current_month_str()
            ws['L1'] = current_month
            
            for age_group, row_num in row_mapping.items():
                self._update_cell_value(ws, f'L{row_num}', negative_pcts[age_group])
            
            # Apply conditional formatting
            #self._apply_conditional_formatting(ws, 'B2:F7', 'white_to_plum')  # Current month data
            #self._apply_conditional_formatting(ws, 'G2:L7', 'white_to_indigo')  # Historical data
        except Exception as e:
            raise Exception(f"Error updating economic future age table: {str(e)}")
    
    def update_econ_current_vs_future_table(self, survey_data: pd.DataFrame):
        """Update the current vs future economic situation table (45_econ_current_vs_future)"""
        try:
            # Find the economy questions
            current_col = self._find_column(survey_data, "Bugün itibari ile ekonominin nasıl olduğunu düşünüyorsunuz")
            future_col = self._find_column(survey_data, "Önümüzdeki bir yıl içerisinde ekonominin nasıl olacağını düşünüyorsunuz")
            
            # Map responses
            survey_data['current_response'] = survey_data[current_col].map(self.econ_current_mapping)
            survey_data['future_response'] = survey_data[future_col].map(self.econ_future_mapping)
            
            # Create pivot table
            pivot_pct = self._create_pivot_table(
                survey_data,
                values='duzeltilmis_agirlik',
                index='current_response',
                columns='future_response',
                calc_method='percent_of_row'
            )
            
            # Get worksheet
            ws = self._get_worksheet('45_econ_current_vs_future')
            
            # Define mappings
            row_mapping = {
                'Çok kötü': 2,
                'Kötü': 3,
                'Ne iyi ne kötü': 4,
                'İyi': 5,
                'Çok iyi': 6
            }
            
            col_mapping = {
                'Çok daha kötü': 'B',
                'Daha kötü': 'C',
                'Değişmez': 'D',
                'Daha iyi': 'E',
                'Çok daha iyi': 'F'
            }
            
            # Update each cell with the new percentage
            for current_response, row_num in row_mapping.items():
                for future_response, col_letter in col_mapping.items():
                    cell_coord = f'{col_letter}{row_num}'
                    try:
                        value = pivot_pct.loc[current_response, future_response]
                        self._update_cell_value(ws, cell_coord, value)
                    except KeyError:
                        self._update_cell_value(ws, cell_coord, 0)
            
            # Apply conditional formatting
            #self._apply_conditional_formatting(ws, 'B2:F6', 'white_to_indigo')
            
        except Exception as e:
            raise Exception(f"Error updating economic current vs future table: {str(e)}")
    
    def update_subsistence_demographics_table(self, survey_data: pd.DataFrame):
        """Update the subsistence by demographics table (50_subsistence_demographics)"""
        try:
            # Find the subsistence question column
            subsistence_col = self._find_column(survey_data, "Aşağıdaki sayılan ifadelerden hangisine katılırsınız")
            
            # Map responses
            response_mapping = {
                'Geçtiğimiz ay gelirim giderlerimi karşılamadı.': 'Karşılamadı',
                'Geçtiğimiz ay gelirim giderlerimi ucu ucuna karşıladı.': 'Ucu ucuna karşıladı',
                'Geçtiğimiz ay gelirim giderlerimin üzerinde oldu.': 'Üzerinde oldu',
                'Geçtiğimiz ay gelirim giderlerimi fazlasıyla karşıladı.': 'Fazlasıyla karşıladı'
            }
            survey_data['subsistence_response'] = survey_data[subsistence_col].map(response_mapping)
            
            # Update age group values
            survey_data['Yaş grubu'] = survey_data['Yaş grubu'].replace('65 ve üstü', '65+')
            
            # Create pivot tables for gender and age
            gender_pivot = self._create_pivot_table(
                survey_data,
                values='duzeltilmis_agirlik',
                index='subsistence_response',
                columns='Katılımcının cinsiyeti?',
                calc_method='percent_of_column'
            )
            
            age_pivot = self._create_pivot_table(
                survey_data,
                values='duzeltilmis_agirlik',
                index='subsistence_response',
                columns='Yaş grubu',
                calc_method='percent_of_column'
            )
            
            # Get worksheet
            ws = self._get_worksheet('50_subsistence_demographics')
            
            # Update gender rows (2-5)
            row_mapping = {
                'Karşılamadı': 2,
                'Ucu ucuna karşıladı': 3,
                'Üzerinde oldu': 4,
                'Fazlasıyla karşıladı': 5
            }
            
            # Update gender columns (B-C)
            gender_col_mapping = {
                'Kadın': 'B',
                'Erkek': 'C'
            }
            
            # Update age columns (D-H)
            age_col_mapping = {
                '18-24': 'D',
                '25-34': 'E',
                '35-44': 'F',
                '45-54': 'G',
                '55-64': 'H',
                '65+': 'I'
            }
            
            # Update gender percentages
            for response, row_num in row_mapping.items():
                for gender, col_letter in gender_col_mapping.items():
                    cell_coord = f'{col_letter}{row_num}'
                    try:
                        value = gender_pivot.loc[response, gender]
                        self._update_cell_value(ws, cell_coord, value)
                    except KeyError:
                        self._update_cell_value(ws, cell_coord, 0)
            
            # Update age percentages
            for response, row_num in row_mapping.items():
                for age_group, col_letter in age_col_mapping.items():
                    cell_coord = f'{col_letter}{row_num}'
                    try:
                        value = age_pivot.loc[response, age_group]
                        self._update_cell_value(ws, cell_coord, value)
                    except KeyError:
                        self._update_cell_value(ws, cell_coord, 0)
            
            # Apply conditional formatting
            #self._apply_conditional_formatting(ws, 'B2:C5', 'white_to_plum')  # Gender groups
            #self._apply_conditional_formatting(ws, 'D2:I5', 'white_to_plum')  # Age groups
        except Exception as e:
            raise Exception(f"Error updating subsistence demographics table: {str(e)}")
    
    def update_subsistence_party_education_table(self, survey_data: pd.DataFrame):
        """Update the subsistence by party and education table (52_subsistence_party_education)"""
        try:
            # Find the subsistence question column
            subsistence_col = self._find_column(survey_data, "Aşağıdaki sayılan ifadelerden hangisine katılırsınız")
            
            # Map responses and party names
            response_mapping = {
                'Geçtiğimiz ay gelirim giderlerimi karşılamadı.': 'Karşılamadı',
                'Geçtiğimiz ay gelirim giderlerimi ucu ucuna karşıladı.': 'Ucu ucuna karşıladı',
                'Geçtiğimiz ay gelirim giderlerimin üzerinde oldu.': 'Üzerinde oldu',
                'Geçtiğimiz ay gelirim giderlerimi fazlasıyla karşıladı.': 'Fazlasıyla karşıladı'
            }
            survey_data['subsistence_response'] = survey_data[subsistence_col].map(response_mapping)
            survey_data['mapped_party'] = survey_data['2023 Genel Seçimlerinde hangi partiye oy verdiniz?'].map(self.party_mapping_2023)
            
            # Create pivot tables for party and education
            party_pivot = self._create_pivot_table(
                survey_data,
                values='duzeltilmis_agirlik',
                index='subsistence_response',
                columns='mapped_party',
                calc_method='percent_of_column'
            )
            
            education_pivot = self._create_pivot_table(
                survey_data,
                values='duzeltilmis_agirlik',
                index='subsistence_response',
                columns='education',
                calc_method='percent_of_column'
            )
            
            # Get worksheet
            ws = self._get_worksheet('52_subsistence_party_education')
            
            # Update response rows (2-5)
            row_mapping = {
                'Karşılamadı': 2,
                'Ucu ucuna karşıladı': 3,
                'Üzerinde oldu': 4,
                'Fazlasıyla karşıladı': 5
            }
            
            # Update party columns (B-F)
            party_col_mapping = {
                'AK Parti': 'B',
                'CHP': 'C',
                'Yeşil Sol Parti': 'D',
                'İYİ Parti': 'E',
                'MHP': 'F'
            }
            
            # Update education columns (G-I)
            education_col_mapping = {
                'İlköğretim ve altı': 'G',
                'Lise': 'H',
                'Yüksekokul ve üzeri': 'I'
            }
            
            # Update party percentages
            for response, row_num in row_mapping.items():
                for party, col_letter in party_col_mapping.items():
                    cell_coord = f'{col_letter}{row_num}'
                    try:
                        value = party_pivot.loc[response, party]
                        self._update_cell_value(ws, cell_coord, value)
                    except KeyError:
                        self._update_cell_value(ws, cell_coord, 0)
            
            # Update education percentages
            for response, row_num in row_mapping.items():
                for edu_level, col_letter in education_col_mapping.items():
                    cell_coord = f'{col_letter}{row_num}'
                    try:
                        value = education_pivot.loc[response, edu_level]
                        self._update_cell_value(ws, cell_coord, value)
                    except KeyError:
                        self._update_cell_value(ws, cell_coord, 0)
            
            # Apply conditional formatting
            #self._apply_conditional_formatting(ws, 'B2:F5', 'white_to_plum')  # Party data
            #self._apply_conditional_formatting(ws, 'G2:I5', 'white_to_indigo')  # Education data
            
        except Exception as e:
            raise Exception(f"Error updating subsistence party education table: {str(e)}")
    
    def update_subsistence_jobs_table(self, survey_data: pd.DataFrame):
        """Update the subsistence by jobs table (53_subsistence_jobs)"""
        try:
            # Find the subsistence question column
            subsistence_col = self._find_column(survey_data, "Aşağıdaki sayılan ifadelerden hangisine katılırsınız")
            
            # Map responses
            response_mapping = {
                'Geçtiğimiz ay gelirim giderlerimi karşılamadı.': 'Karşılamadı',
                'Geçtiğimiz ay gelirim giderlerimi ucu ucuna karşıladı.': 'Ucu ucuna karşıladı',
                'Geçtiğimiz ay gelirim giderlerimin üzerinde oldu.': 'Üzerinde oldu',
                'Geçtiğimiz ay gelirim giderlerimi fazlasıyla karşıladı.': 'Fazlasıyla karşıladı'
            }
            survey_data['subsistence_response'] = survey_data[subsistence_col].map(response_mapping)
            
            # Find jobs column and create pivot table
            jobs_col = self._find_column(survey_data, "Mevcut çalışma durumunuzu belirtir misiniz?")
            survey_data['job_status'] = survey_data[jobs_col]
            
            jobs_pivot = self._create_pivot_table(
                survey_data,
                values='duzeltilmis_agirlik',
                index='subsistence_response',
                columns='job_status',
                calc_method='percent_of_column'
            )
            
            # Get worksheet
            ws = self._get_worksheet('53_subsistence_jobs')
            
            # Update response rows (2-5)
            row_mapping = {
                'Karşılamadı': 2,
                'Ucu ucuna karşıladı': 3,
                'Üzerinde oldu': 4,
                'Fazlasıyla karşıladı': 5
            }
            
            # Update job columns (B-I)
            col_mapping = {
                'Emekli, çalışmıyor': 'B',
                'İşsiz ama iş aramıyor': 'C',
                'İşsiz ve iş arıyor': 'D',
                'Kendi hesabına çalışan veya işveren': 'E',
                'Maaşlı devlet çalışanı': 'F',
                'Öğrenci': 'G',
                'Ücretli özel sektör çalışanı': 'H',
                'Günlük / yevmiyeli çalışan': 'I'
            }
            
            # Update cells with pivot table values
            for response, row_num in row_mapping.items():
                for job, col_letter in col_mapping.items():
                    cell_coord = f'{col_letter}{row_num}'
                    try:
                        value = jobs_pivot.loc[response, job]
                        self._update_cell_value(ws, cell_coord, value)
                    except KeyError:
                        self._update_cell_value(ws, cell_coord, 0)
            
            # Apply conditional formatting
            #self._apply_conditional_formatting(ws, 'B2:I5', 'white_to_indigo')
            
        except Exception as e:
            raise Exception(f"Error updating subsistence jobs table: {str(e)}")
    
    def update_econ_current_jobs_table(self, survey_data: pd.DataFrame):
        """Update the economic situation by jobs table (39_econ_current_jobs)"""
        try:
            # Find the economy question column
            econ_col = self._find_column(survey_data, "Bugün itibari ile ekonominin nasıl olduğunu düşünüyorsunuz")
            
            # Find jobs column and create pivot table
            jobs_col = self._find_column(survey_data, "Mevcut çalışma durumunuzu belirtir misiniz?")
            survey_data['job_status'] = survey_data[jobs_col]
            
            # Map responses
            survey_data['econ_response'] = survey_data[econ_col].map(self.econ_current_mapping)
            
            # Create pivot table
            pivot_pct = self._create_pivot_table(
                survey_data,
                values='duzeltilmis_agirlik',
                index='job_status',
                columns='econ_response',
                calc_method='percent_of_row'
            )
            
            # Get worksheet
            ws = self._get_worksheet('39_econ_current_jobs')
            
            # Update job status rows
            job_row_mapping = {
                'Emekli, çalışmıyor': 2,
                'İşsiz ama iş aramıyor': 3,
                'İşsiz ve iş arıyor': 4,
                'Kendi hesabına çalışan veya işveren': 5,
                'Maaşlı devlet çalışanı': 6,
                'Öğrenci': 7,
                'Ücretli özel sektör çalışanı': 8,
                'Günlük / yevmiyeli çalışan': 9,
            }
            
            col_mapping = {
                'Çok kötü': 'B',
                'Kötü': 'C',
                'Ne iyi ne kötü': 'D',
                'İyi': 'E',
                'Çok iyi': 'F'
            }
            
            # Update job status percentages
            for job, row_num in job_row_mapping.items():
                for response, col_letter in col_mapping.items():
                    cell_coord = f'{col_letter}{row_num}'
                    try:
                        value = pivot_pct.loc[job, response]
                        self._update_cell_value(ws, cell_coord, value)
                    except KeyError:
                        self._update_cell_value(ws, cell_coord, 0)
            
            # Calculate negative percentages for current month
            negative_pcts = {}
            for job in job_row_mapping.keys():
                try:
                    negative_pct = pivot_pct.loc[job, 'Çok kötü'] + pivot_pct.loc[job, 'Kötü']
                    negative_pcts[job] = self._round_values(negative_pct)
                except KeyError:
                    negative_pcts[job] = 0
            
            # Shift historical data left by one column (G to K)
            self._shift_historical_data(ws, 2, 9)
            
            # Update the last column (L) with current month's data
            current_month = self._get_current_month_str()
            ws['G1'] = current_month
            
            for job, row_num in job_row_mapping.items():
                self._update_cell_value(ws, f'G{row_num}', negative_pcts[job])
            
            # Apply conditional formatting
            #self._apply_conditional_formatting(ws, 'B2:F6', 'white_to_plum')  # Current month data
            #self._apply_conditional_formatting(ws, 'G2:G6', 'white_to_indigo')  # Historical data
        except Exception as e:
            raise Exception(f"Error updating economic current jobs table: {str(e)}")
    
    def update_econ_future_jobs_table(self, survey_data: pd.DataFrame):
        """Update the future economic situation by jobs table (45_econ_future_jobs)"""
        try:
            # Find the economy question column
            econ_col = self._find_column(survey_data, "Önümüzdeki bir yıl içerisinde ekonominin nasıl olacağını düşünüyorsunuz")
            
            # Find jobs column and create pivot table
            jobs_col = self._find_column(survey_data, "Mevcut çalışma durumunuzu belirtir misiniz?")
            survey_data['job_status'] = survey_data[jobs_col]
            
            # Map responses
            survey_data['econ_response'] = survey_data[econ_col].map(self.econ_future_mapping)
            
            # Create pivot table
            pivot_pct = self._create_pivot_table(
                survey_data,
                values='duzeltilmis_agirlik',
                index='job_status',
                columns='econ_response',
                calc_method='percent_of_row'
            )
            
            # Get worksheet
            ws = self._get_worksheet('45_econ_future_jobs')
            
            # Update job status rows
            job_row_mapping = {
                'Emekli, çalışmıyor': 2,
                'İşsiz ama iş aramıyor': 3,
                'İşsiz ve iş arıyor': 4,
                'Kendi hesabına çalışan veya işveren': 5,
                'Maaşlı devlet çalışanı': 6,
                'Öğrenci': 7,
                'Ücretli özel sektör çalışanı': 8,
                'Günlük / yevmiyeli çalışan': 9,
            }
            
            col_mapping = {
                'Çok daha kötü': 'B',
                'Daha kötü': 'C',
                'Değişmez': 'D',
                'Daha iyi': 'E',
                'Çok daha iyi': 'F'
            }
            
            # Update job status percentages
            for job, row_num in job_row_mapping.items():
                for response, col_letter in col_mapping.items():
                    cell_coord = f'{col_letter}{row_num}'
                    try:
                        value = pivot_pct.loc[job, response]
                        self._update_cell_value(ws, cell_coord, value)
                    except KeyError:
                        self._update_cell_value(ws, cell_coord, 0)
            
            # Calculate negative percentages for current month
            negative_pcts = {}
            for job in job_row_mapping.keys():
                try:
                    negative_pct = pivot_pct.loc[job, 'Çok daha kötü'] + pivot_pct.loc[job, 'Daha kötü']
                    negative_pcts[job] = self._round_values(negative_pct)
                except KeyError:
                    negative_pcts[job] = 0
            
            # Shift historical data left by one column (G to K)
            self._shift_historical_data(ws, 2, 9)
            
            # Update the last column (L) with current month's data
            current_month = self._get_current_month_str()
            ws['G1'] = current_month
            
            for job, row_num in job_row_mapping.items():
                self._update_cell_value(ws, f'G{row_num}', negative_pcts[job])
            
            # Apply conditional formatting
            #self._apply_conditional_formatting(ws, 'B2:F9', 'white_to_plum')  # Current month data
            #self._apply_conditional_formatting(ws, 'G2:G9', 'white_to_indigo')  # Historical data
        except Exception as e:
            raise Exception(f"Error updating economic future jobs table: {str(e)}")
