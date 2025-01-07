import os
from datetime import datetime
import streamlit as st
import tempfile
from openpyxl import load_workbook

class FileHandler:
    def __init__(self):
        pass
        
    def save_uploaded_file(self, uploaded_file) -> str:
        """Save uploaded file to temp directory and return the path"""
        import tempfile
        import os
        from openpyxl import load_workbook
        
        # Create temp file with same extension
        temp_dir = tempfile.gettempdir()
        file_extension = os.path.splitext(uploaded_file.name)[1]
        temp_path = os.path.join(temp_dir, f'temp_{uploaded_file.name}')
        
        try:
            print(f"Saving uploaded file to: {temp_path}")
            with open(temp_path, 'wb') as f:
                file_content = uploaded_file.getvalue()
                f.write(file_content)
            
            print(f"Successfully saved file, size: {os.path.getsize(temp_path)} bytes")
            
            # Verify the file if it's an Excel file
            if file_extension.lower() == '.xlsx':
                print("Verifying Excel file...")
                try:
                    wb = load_workbook(temp_path)
                    print(f"Excel file verified. Available sheets: {wb.sheetnames}")
                except Exception as e:
                    raise Exception(f"Failed to verify Excel file: {str(e)}")
            
            return temp_path
        except Exception as e:
            print(f"Error saving uploaded file: {str(e)}")
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                    print(f"Cleaned up temporary file: {temp_path}")
                except:
                    pass
            raise
    
    def create_processed_file(self, template_file) -> tuple:
        """Create new files for processing and return paths for both Turkish and English versions"""
        import os
        import tempfile
        
        # Get template path
        template_path = self.save_uploaded_file(template_file)
        
        # Create output paths for both versions
        temp_dir = tempfile.gettempdir()
        name_without_ext, ext = os.path.splitext(template_file.name)
        
        tr_output_filename = f'{name_without_ext}{ext}'
        en_output_filename = f'{name_without_ext}_en{ext}'
        
        tr_output_path = os.path.join(temp_dir, tr_output_filename)
        en_output_path = os.path.join(temp_dir, en_output_filename)
        
        # Copy template to both output paths
        import shutil
        try:
            print(f"Copying template to Turkish version: {tr_output_path}")
            shutil.copy2(template_path, tr_output_path)
            print(f"Successfully copied Turkish version, size: {os.path.getsize(tr_output_path)} bytes")
            
            print(f"Copying template to English version: {en_output_path}")
            shutil.copy2(template_path, en_output_path)
            print(f"Successfully copied English version, size: {os.path.getsize(en_output_path)} bytes")
        except Exception as e:
            print(f"Error copying template: {str(e)}")
            raise
        
        return template_path, tr_output_path, en_output_path
    
    def get_download_button(self, file_path: str, button_text: str = "Download Processed File"):
        """Create a download button for the processed file"""
        try:
            print(f"Creating download button for file: {file_path}")
            
            # Get original filename
            original_name = os.path.basename(file_path)
            
            # Remove 'temp_' prefix if it exists
            if original_name.startswith('temp_'):
                original_name = original_name[5:]
            
            # Only add month suffix if it's not already present
            if not any(month in original_name for month in ['Oca', 'Şub', 'Mar', 'Nis', 'May', 'Haz', 'Tem', 'Ağu', 'Eyl', 'Eki', 'Kas', 'Ara']):
                current_date = datetime.now()
                turkish_months = {
                    1: 'Oca', 2: 'Şub', 3: 'Mar', 4: 'Nis', 5: 'May', 6: 'Haz',
                    7: 'Tem', 8: 'Ağu', 9: 'Eyl', 10: 'Eki', 11: 'Kas', 12: 'Ara'
                }
                month_suffix = f"{turkish_months[current_date.month]}.{str(current_date.year)[2:]}"
                name_without_ext, ext = os.path.splitext(original_name)
                original_name = f"{name_without_ext}_{month_suffix}{ext}"
            
            with open(file_path, 'rb') as file:
                st.download_button(
                    label=button_text,
                    data=file,
                    file_name=original_name,
                    mime="application/vnd.openxmlformats-officedocument.presentationml.presentation"
                )
            print("Successfully created download button")
        except Exception as e:
            print(f"Error creating download button: {str(e)}")
            raise